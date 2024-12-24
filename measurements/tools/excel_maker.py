from openpyxl import load_workbook , Workbook
from openpyxl.styles import Alignment
from os import path

def write_to_excel(excel_dict , particle_samples) :
    file_path = "./measurements/results/result_spreadsheet.xlsx"
    existing_excel = path.exists(file_path)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    if existing_excel : 
        workbook = load_workbook(file_path)
        
    else :
        workbook = Workbook() 
        header_row = [key for key in excel_dict.keys()]
        print("length of sample : " , len(particle_samples))
        header_row.extend([f"particle{i % excel_dict['pop_n']}" for i in range(len(particle_samples))])
    
    val_row = [value for value in excel_dict.values()]
    val_row.extend(particle_samples)

    worksheet = workbook.active

    if not existing_excel : 
        worksheet.append(header_row)

    worksheet.append(val_row)

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(val_row)):
        for cell in row:
            cell.alignment = center_alignment  

    workbook.save(file_path)
